"""Formatted Excel output — three sheets from validated PanelData.

Each sheet carries an analysis metadata row (date/time + source filename)
for traceability.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models.panel import PanelData


@dataclass
class Extraction:
    """One image's extraction result, for use in consolidated Excels."""
    data: PanelData
    analysis_datetime: datetime
    source_filename: str

logger = logging.getLogger(__name__)

# ── Styles ─────────────────────────────────────────────────────────────

_THIN = Side(style="thin")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="2F5496")
META_FONT = Font(name="Calibri", italic=True, size=10, color="555555")

TITLE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
BLUE_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
ALT_ROW = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HIGHLIGHT = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
META_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


# ── Cell helpers ───────────────────────────────────────────────────────

def write_cell(ws: Worksheet, row: int, col: int, value: object,
               font: Font = DATA_FONT, fill: PatternFill | None = None,
               alignment: Alignment = CENTER,
               border: Border | None = BORDER) -> None:
    """Write a single styled cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = alignment
    if border:
        cell.border = border
    if fill:
        cell.fill = fill


def write_title(ws: Worksheet, row: int, text: str, col_start: int, col_end: int,
                fill: PatternFill = TITLE_FILL) -> None:
    """Write a merged title row spanning multiple columns."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    write_cell(ws, row, col_start, text, TITLE_FONT, fill)
    for c in range(col_start + 1, col_end + 1):
        write_cell(ws, row, c, None, TITLE_FONT, fill)
    ws.row_dimensions[row].height = 30


def write_metadata(ws: Worksheet, row: int, analysis_dt: datetime,
                   source_filename: str | None, col_end: int) -> None:
    """Write a metadata row with analysis date/time and source filename."""
    date_str = analysis_dt.strftime("%d/%m/%Y a %H:%M:%S")
    if source_filename:
        text = f"Analyse effectuee le {date_str}  |  Source: {source_filename}"
    else:
        text = f"Analyse effectuee le {date_str}"

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = META_FONT
    cell.alignment = CENTER
    cell.fill = META_FILL
    ws.row_dimensions[row].height = 22


def write_headers(ws: Worksheet, row: int, headers: list[str],
                  fill: PatternFill = BLUE_HEADER) -> None:
    """Write a header row."""
    for i, h in enumerate(headers, 1):
        write_cell(ws, row, i, h, HEADER_FONT, fill)
    ws.row_dimensions[row].height = 28


def set_widths(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet builders ─────────────────────────────────────────────────────

def _build_puissance(ws: Worksheet, data: PanelData, analysis_dt: datetime,
                     source: str | None) -> None:
    write_title(ws, 1, "PUISSANCE ICT — Electron Processing System AQUILLA", 1, 6)
    write_metadata(ws, 2, analysis_dt, source, col_end=6)
    write_headers(ws, 4, [
        "Zone", "Tension primaire (V)", "Zone Cp",
        "Courant primaire (A)", "Zone Cs", "Courant secondaire (A)",
    ])
    for i, r in enumerate(data.puissance_ict):
        row = 5 + i
        fill = ALT_ROW if i % 2 == 0 else None
        write_cell(ws, row, 1, r.zone_tension, BOLD_FONT, fill)
        write_cell(ws, row, 2, r.tension_primaire_v, DATA_FONT, fill)
        write_cell(ws, row, 3, r.zone_courant_primaire, BOLD_FONT, fill)
        write_cell(ws, row, 4, r.courant_primaire_a, DATA_FONT, fill)
        write_cell(ws, row, 5, r.zone_courant_secondaire, BOLD_FONT, fill)
        write_cell(ws, row, 6, r.courant_secondaire_a, DATA_FONT, fill)
    set_widths(ws, [12, 24, 12, 24, 12, 26])


def _build_accelerateurs(ws: Worksheet, data: PanelData, analysis_dt: datetime,
                         source: str | None) -> None:
    write_title(ws, 1, "ACCELERATEURS — Electron Processing System AQUILLA", 1, 5)
    write_metadata(ws, 2, analysis_dt, source, col_end=5)

    ws.merge_cells("A4:E4")
    write_cell(ws, 4, 1, "Parametres par Zone", SECTION_FONT, alignment=LEFT)
    write_headers(ws, 5, [
        "Zone", "R Icol (KV)", "Courant colonne (uA)", "Vide (Torr)", "Courant aperture (uA)",
    ], GREEN_HEADER)
    for i, r in enumerate(data.accelerateurs_zones):
        row = 6 + i
        fill = ALT_ROW if i % 2 == 0 else None
        write_cell(ws, row, 1, r.zone, BOLD_FONT, fill)
        write_cell(ws, row, 2, r.r_icol_kv, DATA_FONT, fill)
        write_cell(ws, row, 3, r.courant_colonne_ua, DATA_FONT, fill)
        write_cell(ws, row, 4, r.vide_torr, DATA_FONT, fill)
        write_cell(ws, row, 5, r.courant_aperture_ua, DATA_FONT, fill)

    ws.merge_cells("A9:E9")
    write_cell(ws, 9, 1, "Parametres Globaux", SECTION_FONT, alignment=LEFT)
    write_headers(ws, 10, ["Parametre", "Valeur", "Unite", "", ""], GREEN_HEADER)
    g = data.accelerateurs_global
    for i, (p, v, u) in enumerate([
        ("Tension KV", g.tension_kv, "KV"), ("Charge mA", g.charge_ma, "mA"),
        ("Faisceau mA - Zone A", g.faisceau_ma_a, "mA"),
        ("Faisceau mA - Zone B", g.faisceau_ma_b, "mA"),
    ]):
        row = 11 + i
        fill = HIGHLIGHT if "Faisceau" in p else (ALT_ROW if i % 2 == 0 else None)
        write_cell(ws, row, 1, p, DATA_FONT, fill)
        write_cell(ws, row, 2, v, DATA_FONT, fill)
        write_cell(ws, row, 3, u, DATA_FONT, fill)
    set_widths(ws, [28, 22, 25, 18, 25])


def _build_resume(ws: Worksheet, data: PanelData, analysis_dt: datetime,
                  source: str | None) -> None:
    write_title(ws, 1, "RESUME COMPLET — Panel Electronique AQUILLA", 1, 7)
    write_metadata(ws, 2, analysis_dt, source, col_end=7)

    write_title(ws, 4, "PUISSANCE ICT", 1, 7, BLUE_HEADER)
    write_headers(ws, 5, ["Zone", "Tension (V)", "Zp", "Courant prim (A)", "Zs", "Courant sec (A)", ""])
    for i, r in enumerate(data.puissance_ict):
        row = 6 + i
        fill = ALT_ROW if i % 2 == 0 else None
        for j, v in enumerate([r.zone_tension, r.tension_primaire_v, r.zone_courant_primaire,
                                r.courant_primaire_a, r.zone_courant_secondaire,
                                r.courant_secondaire_a, None], 1):
            write_cell(ws, row, j, v, BOLD_FONT if j in (1, 3, 5) else DATA_FONT, fill)

    write_title(ws, 10, "ACCELERATEURS", 1, 7, GREEN_HEADER)
    write_headers(ws, 11, ["Zone", "R Icol (KV)", "Courant (uA)", "Vide (Torr)", "Aperture (uA)", "", ""], GREEN_HEADER)
    for i, r in enumerate(data.accelerateurs_zones):
        row = 12 + i
        fill = ALT_ROW if i % 2 == 0 else None
        for j, v in enumerate([r.zone, r.r_icol_kv, r.courant_colonne_ua,
                                r.vide_torr, r.courant_aperture_ua, None, None], 1):
            write_cell(ws, row, j, v, BOLD_FONT if j == 1 else DATA_FONT, fill)

    ws.merge_cells("A15:C15")
    write_cell(ws, 15, 1, "VALEURS GLOBALES", Font(bold=True, size=11, color="548235"), alignment=LEFT)
    write_headers(ws, 16, ["Parametre", "Valeur", "Unite"], GREEN_HEADER)
    g = data.accelerateurs_global
    for i, (p, v, u) in enumerate([
        ("Tension KV", g.tension_kv, "KV"), ("Charge mA", g.charge_ma, "mA"),
        ("Faisceau A", g.faisceau_ma_a, "mA"), ("Faisceau B", g.faisceau_ma_b, "mA"),
    ]):
        write_cell(ws, 17 + i, 1, p, DATA_FONT)
        write_cell(ws, 17 + i, 2, v, DATA_FONT)
        write_cell(ws, 17 + i, 3, u, DATA_FONT)
    set_widths(ws, [22, 22, 22, 22, 22, 14, 14])


# ── Public API ─────────────────────────────────────────────────────────

def generate_excel(
    data: PanelData,
    output_path: str | Path,
    analysis_datetime: datetime | None = None,
    source_filename: str | None = None,
) -> Path:
    """Write validated PanelData to a formatted .xlsx file.

    Args:
        data: Validated panel data.
        output_path: Where to save the .xlsx file.
        analysis_datetime: When the analysis was performed. Defaults to now.
        source_filename: Original screenshot filename (for traceability).

    Returns:
        The path to the saved file.
    """
    if analysis_datetime is None:
        analysis_datetime = datetime.now()

    out = Path(output_path)
    wb = Workbook()

    _build_puissance(wb.active, data, analysis_datetime, source_filename)
    wb.active.title = "Puissance ICT"
    _build_accelerateurs(wb.create_sheet("Accelerateurs"), data, analysis_datetime, source_filename)
    _build_resume(wb.create_sheet("Resume Complet"), data, analysis_datetime, source_filename)

    # Workbook-level metadata
    wb.properties.creator = "Panel Extractor"
    wb.properties.title = f"AQUILLA Panel Data — {analysis_datetime.strftime('%Y-%m-%d %H:%M')}"
    wb.properties.created = analysis_datetime
    wb.properties.modified = analysis_datetime

    wb.save(str(out))
    logger.info(f"Excel saved: {out} ({out.stat().st_size / 1024:.1f} KB)")
    return out


# ═══════════════════════════════════════════════════════════════════════════
# Consolidated Excel (multiple images in ONE file)
# ═══════════════════════════════════════════════════════════════════════════

def _fmt_dt(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y %H:%M:%S")


def _build_consolidated_puissance(ws: Worksheet, extractions: list[Extraction]) -> None:
    """Puissance ICT sheet — 3 rows per image, stacked."""
    n = len(extractions)
    write_title(ws, 1, f"PUISSANCE ICT — {n} extraction(s) consolidees", 1, 8)

    subtitle = f"Genere le {_fmt_dt(datetime.now())}  |  {n} image(s)"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.font = META_FONT
    cell.alignment = CENTER
    cell.fill = META_FILL
    ws.row_dimensions[2].height = 22

    write_headers(ws, 4, [
        "Source", "Date analyse", "Zone",
        "Tension primaire (V)", "Zone Cp", "Courant primaire (A)",
        "Zone Cs", "Courant secondaire (A)",
    ])

    row = 5
    for i, ext in enumerate(extractions):
        # Same image = same fill (visual grouping)
        fill = ALT_ROW if i % 2 == 0 else None
        date_str = _fmt_dt(ext.analysis_datetime)
        for r in ext.data.puissance_ict:
            write_cell(ws, row, 1, ext.source_filename, DATA_FONT, fill)
            write_cell(ws, row, 2, date_str, DATA_FONT, fill)
            write_cell(ws, row, 3, r.zone_tension, BOLD_FONT, fill)
            write_cell(ws, row, 4, r.tension_primaire_v, DATA_FONT, fill)
            write_cell(ws, row, 5, r.zone_courant_primaire, BOLD_FONT, fill)
            write_cell(ws, row, 6, r.courant_primaire_a, DATA_FONT, fill)
            write_cell(ws, row, 7, r.zone_courant_secondaire, BOLD_FONT, fill)
            write_cell(ws, row, 8, r.courant_secondaire_a, DATA_FONT, fill)
            row += 1

    set_widths(ws, [22, 20, 10, 22, 12, 22, 12, 22])


def _build_consolidated_accelerateurs(ws: Worksheet, extractions: list[Extraction]) -> None:
    """Accelerateurs sheet — 2 zone rows per image, stacked."""
    n = len(extractions)
    write_title(ws, 1, f"ACCELERATEURS — {n} extraction(s) consolidees", 1, 7)

    subtitle = f"Genere le {_fmt_dt(datetime.now())}  |  {n} image(s)"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.font = META_FONT
    cell.alignment = CENTER
    cell.fill = META_FILL
    ws.row_dimensions[2].height = 22

    write_headers(ws, 4, [
        "Source", "Date analyse", "Zone",
        "R Icol (KV)", "Courant colonne (uA)", "Vide (Torr)", "Courant aperture (uA)",
    ], GREEN_HEADER)

    row = 5
    for i, ext in enumerate(extractions):
        fill = ALT_ROW if i % 2 == 0 else None
        date_str = _fmt_dt(ext.analysis_datetime)
        for r in ext.data.accelerateurs_zones:
            write_cell(ws, row, 1, ext.source_filename, DATA_FONT, fill)
            write_cell(ws, row, 2, date_str, DATA_FONT, fill)
            write_cell(ws, row, 3, r.zone, BOLD_FONT, fill)
            write_cell(ws, row, 4, r.r_icol_kv, DATA_FONT, fill)
            write_cell(ws, row, 5, r.courant_colonne_ua, DATA_FONT, fill)
            write_cell(ws, row, 6, r.vide_torr, DATA_FONT, fill)
            write_cell(ws, row, 7, r.courant_aperture_ua, DATA_FONT, fill)
            row += 1

    set_widths(ws, [22, 20, 10, 14, 22, 18, 22])


def _build_consolidated_global(ws: Worksheet, extractions: list[Extraction]) -> None:
    """Valeurs Globales sheet — 1 row per image."""
    n = len(extractions)
    write_title(ws, 1, f"VALEURS GLOBALES — {n} extraction(s) consolidees", 1, 6, GREEN_HEADER)

    subtitle = f"Genere le {_fmt_dt(datetime.now())}  |  {n} image(s)"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.font = META_FONT
    cell.alignment = CENTER
    cell.fill = META_FILL
    ws.row_dimensions[2].height = 22

    write_headers(ws, 4, [
        "Source", "Date analyse",
        "Tension (KV)", "Charge (mA)", "Faisceau A (mA)", "Faisceau B (mA)",
    ], GREEN_HEADER)

    for i, ext in enumerate(extractions):
        row = 5 + i
        fill = ALT_ROW if i % 2 == 0 else None
        g = ext.data.accelerateurs_global
        write_cell(ws, row, 1, ext.source_filename, DATA_FONT, fill)
        write_cell(ws, row, 2, _fmt_dt(ext.analysis_datetime), DATA_FONT, fill)
        write_cell(ws, row, 3, g.tension_kv, DATA_FONT, fill)
        write_cell(ws, row, 4, g.charge_ma, DATA_FONT, fill)
        write_cell(ws, row, 5, g.faisceau_ma_a, DATA_FONT, fill)
        write_cell(ws, row, 6, g.faisceau_ma_b, DATA_FONT, fill)

    set_widths(ws, [22, 20, 16, 16, 18, 18])


def _build_consolidated_resume(ws: Worksheet, extractions: list[Extraction]) -> None:
    """Complete resume — ALL 21 values per extraction on one row.

    Column layout (25 cols):
      [Metadata: 2] Source, Date
      [Puissance ICT B-C: 3] Tension, Courant prim, Courant sec
      [Puissance ICT C-A: 3] Tension, Courant prim, Courant sec
      [Puissance ICT A-B: 3] Tension, Courant prim, Courant sec
      [Accelerateur A: 4] R Icol, Courant colonne, Vide, Aperture
      [Accelerateur B: 4] R Icol, Courant colonne, Vide, Aperture
      [Global: 4] Tension KV, Charge mA, Faisceau A, Faisceau B
    """
    n = len(extractions)
    NUM_COLS = 25

    write_title(ws, 1, f"RESUME COMPLET — {n} extraction(s)", 1, NUM_COLS)

    subtitle = f"Toutes les valeurs (21 par extraction)  |  Genere le {_fmt_dt(datetime.now())}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.font = META_FONT
    cell.alignment = CENTER
    cell.fill = META_FILL
    ws.row_dimensions[2].height = 22

    write_headers(ws, 4, [
        "Source", "Date",
        # Puissance ICT
        "B-C Tension (V)", "B-C Ip A (A)", "B-C Is A (A)",
        "C-A Tension (V)", "C-A Ip B (A)", "C-A Is B (A)",
        "A-B Tension (V)", "A-B Ip C (A)", "A-B Is C (A)",
        # Accelerateurs
        "A R Icol (KV)", "A Courant col (uA)", "A Vide (Torr)", "A Aperture (uA)",
        "B R Icol (KV)", "B Courant col (uA)", "B Vide (Torr)", "B Aperture (uA)",
        # Global
        "Tension (KV)", "Charge (mA)", "Faisceau A (mA)", "Faisceau B (mA)",
    ])

    def _puiss_row(rows: list, zone: str):
        for r in rows:
            if r.zone_tension == zone:
                return r
        return None

    def _accel_row(rows: list, zone: str):
        for r in rows:
            if r.zone == zone:
                return r
        return None

    for i, ext in enumerate(extractions):
        row = 5 + i
        fill = ALT_ROW if i % 2 == 0 else None
        d = ext.data
        g = d.accelerateurs_global

        p_bc = _puiss_row(d.puissance_ict, "B-C")
        p_ca = _puiss_row(d.puissance_ict, "C-A")
        p_ab = _puiss_row(d.puissance_ict, "A-B")
        a_a = _accel_row(d.accelerateurs_zones, "A")
        a_b = _accel_row(d.accelerateurs_zones, "B")

        cols = [
            ext.source_filename,
            _fmt_dt(ext.analysis_datetime),
            # Puissance ICT
            p_bc.tension_primaire_v, p_bc.courant_primaire_a, p_bc.courant_secondaire_a,
            p_ca.tension_primaire_v, p_ca.courant_primaire_a, p_ca.courant_secondaire_a,
            p_ab.tension_primaire_v, p_ab.courant_primaire_a, p_ab.courant_secondaire_a,
            # Accelerateurs
            a_a.r_icol_kv, a_a.courant_colonne_ua, a_a.vide_torr, a_a.courant_aperture_ua,
            a_b.r_icol_kv, a_b.courant_colonne_ua, a_b.vide_torr, a_b.courant_aperture_ua,
            # Global
            g.tension_kv, g.charge_ma, g.faisceau_ma_a, g.faisceau_ma_b,
        ]
        for j, v in enumerate(cols, 1):
            write_cell(ws, row, j, v, DATA_FONT, fill)

    set_widths(ws, [
        22, 18,
        13, 11, 11,
        13, 11, 11,
        13, 11, 11,
        12, 16, 14, 14,
        12, 16, 14, 14,
        12, 12, 14, 14,
    ])
    # Freeze panes: keep source + date visible when scrolling right
    ws.freeze_panes = "C5"


def generate_consolidated_excel(
    extractions: list[Extraction],
    output_path: str | Path,
) -> Path:
    """Build ONE Excel file with data from multiple images consolidated.

    Sheets:
      1. Puissance ICT       — 3 rows per image (stacked)
      2. Accelerateurs       — 2 zone rows per image (stacked)
      3. Valeurs Globales    — 1 row per image
      4. Resume Complet      — 1 row per image with key values

    Every data row has Source + Date columns for full traceability.
    """
    if not extractions:
        raise ValueError("No extractions to consolidate")

    out = Path(output_path)
    wb = Workbook()

    _build_consolidated_puissance(wb.active, extractions)
    wb.active.title = "Puissance ICT"
    _build_consolidated_accelerateurs(wb.create_sheet("Accelerateurs"), extractions)
    _build_consolidated_global(wb.create_sheet("Valeurs Globales"), extractions)
    _build_consolidated_resume(wb.create_sheet("Resume Complet"), extractions)

    # Workbook metadata
    wb.properties.creator = "Panel Extractor"
    wb.properties.title = f"AQUILLA Panel Data (consolidated) — {len(extractions)} images"
    now = datetime.now()
    wb.properties.created = now
    wb.properties.modified = now

    wb.save(str(out))
    logger.info(f"Consolidated Excel saved: {out} ({len(extractions)} images, {out.stat().st_size / 1024:.1f} KB)")
    return out
